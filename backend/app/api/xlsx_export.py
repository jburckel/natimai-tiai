"""Excel responses for the console's export buttons.

The CSV export exists because Excel opens it; this one exists because Excel
opens it *well*: a real date is sortable and filterable, a real number sums,
and a header that is bold, frozen and carrying an auto-filter is what a
spreadsheet handed to a meeting looks like. The column set is the caller's
business — this module only knows how to lay a header and rows out.
"""

import io
from collections.abc import Iterable, Sequence
from datetime import date, datetime

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Widest a column is allowed to grow when sized to its content: a 300-character
# CPU string or install path would otherwise push every other column off the
# screen. Narrowest, so a header of two letters still reads.
MIN_WIDTH = 8
MAX_WIDTH = 60

DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
DATE_FORMAT = "yyyy-mm-dd"


def xlsx_response(
    filename: str,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
    *,
    sheet_title: str = "Export",
) -> Response:
    """Render rows as a downloadable ``.xlsx`` workbook with one sheet.

    ``None`` becomes an empty cell, as in the CSV: a spreadsheet has no notion
    of "never reported". Dates and datetimes are written as Excel dates with a
    display format rather than as text, which is the whole reason to prefer
    this format over CSV. Aware datetimes must be made naive by the caller —
    Excel has no timezone, and openpyxl refuses an aware value.
    """
    workbook = Workbook()
    # A new sheet rather than the default one: ``active`` is typed as possibly
    # absent, and the default sheet is dropped so the workbook opens on ours.
    sheet = workbook.create_sheet(title=sheet_title, index=0)
    for other in list(workbook.worksheets):
        if other is not sheet:
            workbook.remove(other)

    sheet.append(list(header))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    widths = [len(h) for h in header]
    for row in rows:
        values = ["" if v is None else v for v in row]
        sheet.append(values)
        for i, value in enumerate(values):
            if isinstance(value, datetime):
                sheet.cell(
                    row=sheet.max_row, column=i + 1
                ).number_format = DATETIME_FORMAT
                length = len(DATETIME_FORMAT)
            elif isinstance(value, date):
                sheet.cell(row=sheet.max_row, column=i + 1).number_format = DATE_FORMAT
                length = len(DATE_FORMAT)
            else:
                length = len(str(value))
            if length > widths[i]:
                widths[i] = length

    for i, width in enumerate(widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = min(
            MAX_WIDTH, max(MIN_WIDTH, width + 2)
        )
    # The header stays put while the rows scroll, and every column gets the
    # dropdown Excel users reach for first.
    sheet.freeze_panes = "A2"
    if header:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(header))}{max(sheet.max_row, 1)}"
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
