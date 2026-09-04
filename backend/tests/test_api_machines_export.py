"""The fleet export with chosen columns, in CSV and Excel (require TIAI_TEST_DATABASE_URL).

Also covers the inventory facets added with it — processor, chassis kind,
memory bounds — and the two listings that feed their dropdowns.
"""

import io

from openpyxl import load_workbook

from tests.test_api_inventory import (  # noqa: F401  (fixtures come from conftest)
    _enroll,
    _heartbeat,
    _inventory,
)
from tests.test_api_inventory_fleet import _admin_headers, _ids, _poste

# --- Column catalogue --------------------------------------------------------


async def test_export_columns_lists_the_catalogue(client, db_session):
    """The picker reads the catalogue from the server: keys, labels, groups, defaults."""
    headers = await _admin_headers(client, db_session)
    resp = await client.get("/api/v1/machines/export-columns", headers=headers)
    assert resp.status_code == 200, resp.text
    columns = resp.json()
    keys = [c["key"] for c in columns]
    assert "hostname" in keys and "cpu_model" in keys and "ram_total_gb" in keys
    by_key = {c["key"]: c for c in columns}
    assert by_key["hostname"]["default"] is True
    assert by_key["mac_address"]["default"] is False
    assert by_key["cpu_model"]["group"] == "hardware"
    assert by_key["cpu_model"]["group_label"] == "Matériel"
    assert by_key["last_seen"]["kind"] == "datetime"


# --- CSV ----------------------------------------------------------------------


async def test_csv_export_defaults_to_the_usual_columns(client, db_session):
    """No ``columns``: the default set, in catalogue order."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "exp-default", hostname="PC-DEF")

    resp = await client.get("/api/v1/machines/export.csv", headers=headers)
    assert resp.status_code == 200, resp.text
    first_line = resp.text.lstrip("﻿").splitlines()[0]
    assert first_line.startswith("Nom;Domaine;Adresse IP;OS;Architecture")
    assert "Processeur" in first_line
    assert "RAM (Gio)" in first_line
    assert "Adresse MAC" not in first_line


async def test_csv_export_honours_chosen_columns_and_order(client, db_session):
    """The columns asked for, in the order asked — nothing more."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "exp-cols", hostname="PC-COLS")

    resp = await client.get(
        "/api/v1/machines/export.csv?columns=cpu_cores,hostname,ram_total_gb,hw_is_virtual",
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    lines = resp.text.lstrip("﻿").splitlines()
    assert lines[0] == "Cœurs;Nom;RAM (Gio);Machine virtuelle"
    assert lines[1] == "16;PC-COLS;32;Non"


async def test_csv_export_rejects_an_unknown_column(client, db_session):
    """A silently shorter file would only be noticed in the meeting."""
    headers = await _admin_headers(client, db_session)
    resp = await client.get(
        "/api/v1/machines/export.csv?columns=hostname,password_hash", headers=headers
    )
    assert resp.status_code == 422, resp.text
    assert resp.json()["error"]["details"]["unknown"] == ["password_hash"]


async def test_csv_export_writes_timestamps_in_the_readers_zone(client, db_session):
    """« Vu le » must read as the console showed it, not two hours off."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "exp-tz", hostname="PC-TZ")

    utc = await client.get(
        "/api/v1/machines/export.csv?columns=last_seen", headers=headers
    )
    paris = await client.get(
        "/api/v1/machines/export.csv?columns=last_seen&tz=Europe/Paris",
        headers=headers,
    )
    utc_value = utc.text.lstrip("﻿").splitlines()[1]
    paris_value = paris.text.lstrip("﻿").splitlines()[1]
    assert utc_value != paris_value
    # Neither is a raw ISO dump with seconds and a zone suffix.
    assert len(utc_value) == len("2026-09-04 09:58")

    # An unknown zone falls back on UTC rather than failing the export.
    bogus = await client.get(
        "/api/v1/machines/export.csv?columns=last_seen&tz=Mars/Olympus",
        headers=headers,
    )
    assert bogus.status_code == 200
    assert bogus.text.lstrip("﻿").splitlines()[1] == utc_value


# --- Excel --------------------------------------------------------------------


async def test_xlsx_export_is_a_real_workbook(client, db_session):
    """Header, values, and dates as dates: what makes it worth more than CSV."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "exp-xlsx", hostname="PC-XLSX")

    resp = await client.get(
        "/api/v1/machines/export.xlsx?columns=hostname,ram_total_gb,bios_date,last_seen,secure_boot",
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "parc.xlsx" in resp.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(resp.content))
    sheet = workbook["Parc"]
    assert [c.value for c in sheet[1]] == [
        "Nom",
        "RAM (Gio)",
        "Date BIOS",
        "Vu le",
        "Secure Boot",
    ]
    row = [c.value for c in sheet[2]]
    assert row[0] == "PC-XLSX"
    assert row[1] == 32
    # Real dates, not strings — the reason to prefer this format.
    assert row[2].isoformat().startswith("2024-01-15")
    assert row[3].tzinfo is None
    assert row[4] == "Oui"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


async def test_xlsx_export_honours_the_filters(client, db_session):
    """Same facets as the list: an export of another list is worse than none."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "exp-f1", hostname="PC-DELL", hw_model="OptiPlex 7010")
    await _poste(client, "exp-f2", hostname="PC-LENOVO", hw_model="ThinkPad T14")

    resp = await client.get(
        "/api/v1/machines/export.xlsx?columns=hostname&hw_model=ThinkPad",
        headers=headers,
    )
    sheet = load_workbook(io.BytesIO(resp.content))["Parc"]
    names = [row[0].value for row in sheet.iter_rows(min_row=2)]
    assert names == ["PC-LENOVO"]


async def test_exports_are_forbidden_for_the_unauthenticated(client, db_session):
    for path in (
        "/api/v1/machines/export.xlsx",
        "/api/v1/machines/export-columns",
    ):
        resp = await client.get(path)
        assert resp.status_code == 401, path


# --- The new inventory facets ------------------------------------------------


async def test_filter_by_processor_and_chassis(client, db_session):
    """Processor as a substring (a generation), chassis as an exact kind."""
    headers = await _admin_headers(client, db_session)
    i7 = await _poste(
        client,
        "fac-i7",
        cpu_model="Intel(R) Core(TM) i7-13700",
        hw_chassis_type="desktop",
    )
    i5 = await _poste(
        client,
        "fac-i5",
        cpu_model="Intel(R) Core(TM) i5-8250U",
        hw_chassis_type="laptop",
    )

    resp = await client.get("/api/v1/machines?cpu_model=i5-8", headers=headers)
    assert await _ids(resp) == {i5["machine_id"]}

    resp = await client.get("/api/v1/machines?hw_chassis_type=desktop", headers=headers)
    assert await _ids(resp) == {i7["machine_id"]}


async def test_filter_by_memory_bounds_in_nominal_gib(client, db_session):
    """16 GiB reported as 16 289 MiB must still count as 16 — and 8 GiB as 8."""
    headers = await _admin_headers(client, db_session)
    sixteen = await _poste(client, "ram-16", ram_total_mb=16_289)
    eight = await _poste(client, "ram-8", ram_total_mb=8_100)
    thirty_two = await _poste(client, "ram-32", ram_total_mb=32_768)
    # Never reported: neither small nor large.
    silent = await _enroll(client, "ram-silent")

    at_least_16 = await client.get("/api/v1/machines?ram_min_gb=16", headers=headers)
    assert await _ids(at_least_16) == {sixteen["machine_id"], thirty_two["machine_id"]}

    at_most_8 = await client.get("/api/v1/machines?ram_max_gb=8", headers=headers)
    assert await _ids(at_most_8) == {eight["machine_id"]}

    between = await client.get(
        "/api/v1/machines?ram_min_gb=8&ram_max_gb=16", headers=headers
    )
    assert await _ids(between) == {eight["machine_id"], sixteen["machine_id"]}
    assert silent["machine_id"] not in await _ids(between)


async def test_processors_and_chassis_listings(client, db_session):
    """Fleet data, most widespread first — the shape /models established."""
    headers = await _admin_headers(client, db_session)
    await _poste(client, "lst-1", cpu_model="Intel i5-8250U", hw_chassis_type="laptop")
    await _poste(client, "lst-2", cpu_model="Intel i5-8250U", hw_chassis_type="laptop")
    await _poste(client, "lst-3", cpu_model="AMD Ryzen 5", hw_chassis_type="desktop")

    cpus = await client.get("/api/v1/machines/processors", headers=headers)
    assert cpus.status_code == 200, cpus.text
    assert cpus.json()[0] == {"name": "Intel i5-8250U", "count": 2}

    kinds = await client.get("/api/v1/machines/chassis-types", headers=headers)
    assert kinds.json()[0] == {"name": "laptop", "count": 2}
